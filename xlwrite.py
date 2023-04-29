#import xlwt;
from datetime import datetime;
#from xlutils.copy import copy
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color

'''style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
    num_format_str='#,##0.00')
style1 = xlwt.easyxf(num_format_str='D-MMM-YY')

wb = xlwt.Workbook()
ws = wb.add_sheet('A Test Sheet')

ws.write(0, 0, 1234.56, style0)
ws.write(1, 0, datetime.now(), style1)
ws.write(2, 0, 1)
ws.write(2, 1, 1)
ws.write(2, 2, xlwt.Formula("A3+B3"))

wb.save('example.xls')
'''
def output(filename, sheet, name, roll,tim):
    my_file = Path('C:/Users/panki/Desktop/attendance/attendance_sheets/'+filename+str(datetime.now().date())+'.xlsx');
    if my_file.is_file():
        book = openpyxl.load_workbook('C:/Users/panki/Desktop/attendance/attendance_sheets/'+filename+str(datetime.now().date())+'.xlsx');
        sh = book.active
        # file exists
    else:
        book = Workbook()
        sh = book.active

    #variables = [x, y, z]
    #x_desc = 'Display'
    #y_desc = 'Dominance'
    #z_desc = 'Test'
    #desc = [x_desc, y_desc, z_desc]
    #sh.write(0,0,datetime.now().date(),style1)

    ft = Font(color="FF0000",bold=True)
    col1_name = 'Name'
    col2_name = 'Roll No'
    col3_name = 'Time'


    sh.cell(1,1,col1_name)
    sh.cell(1, 2, col2_name)
    sh.cell(1, 3, col3_name)

    a = sh['A1']
    a.font = ft
    a = sh['B1']
    a.font = ft
    a = sh['C1']
    a.font = ft

    row = sh.max_row

    sh.cell(row+1,1,name);
    sh.cell(row+1, 2, roll);
    sh.cell(row + 1, 3, tim);
    #You may need to group the variables together
    #for n, (v_desc, v) in enumerate(zip(desc, variables)):
    fullname=filename+str(datetime.now().date())+'.xlsx';
    book.save('C:/Users/panki/Desktop/attendance/attendance_sheets/'+fullname)