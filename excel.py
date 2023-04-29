from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.cell(1,1,'name')
ws.cell(1,2,'Time')

wb.save('demo.xlsx')